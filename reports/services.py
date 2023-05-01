# --------------------------------------------------------------------------------------------------
#  COPYRIGHT(c) 2023 MONTA                                                                         -
#                                                                                                  -
#  This file is part of Monta.                                                                     -
#                                                                                                  -
#  The Monta software is licensed under the Business Source License 1.1. You are granted the right -
#  to copy, modify, and redistribute the software, but only for non-production use or with a total -
#  of less than three server instances. Starting from the Change Date (November 16, 2026), the     -
#  software will be made available under version 2 or later of the GNU General Public License.     -
#  If you use the software in violation of this license, your rights under the license will be     -
#  terminated automatically. The software is provided "as is," and the Licensor disclaims all      -
#  warranties and conditions. If you use this license's text or the "Business Source License" name -
#  and trademark, you must comply with the Licensor's covenants, which include specifying the      -
#  Change License as the GPL Version 2.0 or a compatible license, specifying an Additional Use     -
#  Grant, and not modifying the license in any other way.                                          -
# --------------------------------------------------------------------------------------------------

import io
import json

from django.apps import apps
from django.db.models import Model
from django_celery_beat.models import CrontabSchedule, PeriodicTask
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reports import exceptions, models


def get_model_by_table_name(table_name: str) -> type[Model] | None:
    """
    Returns a model class from Django apps by a given table name.
    If the model class is not found, it returns None.

    Args:
        table_name (str): A string representing the name of the database table.

    Returns:
        Type[models.Model]: A model class that corresponds to the given table name.
    """
    return next(
        (model for model in apps.get_models() if model._meta.db_table == table_name),
        None,
    )


def generate_excel_report_as_file(report: models.CustomReport) -> io.BytesIO:
    """Generate an Excel report as a file.

    Args:
        report (CustomReport): A Report object that specifies the table and columns to use for the report.

    Returns:
        io.BytesIO: A BytesIO object that contains the generated Excel file.

    This function takes a Report object and generates an Excel file based on the specified
    table and columns.
    The function uses openpyxl library to create a new workbook, writes the headers of the
    columns, and populates the data rows by iterating through the data model object.
    The generated Excel file is saved to a BytesIO object which is then returned to the caller.
    """
    model: type[Model] | type[Model] | None = get_model_by_table_name(report.table)

    if not model:
        raise exceptions.InvalidTableException("Invalid table name.")

    columns = report.columns.all().order_by("column_order")

    wb = Workbook()
    ws = wb.active

    for index, column in enumerate(columns):
        col_letter = get_column_letter(index + 1)
        ws[f"{col_letter}1"] = column.column_name

    for row, obj in enumerate(model.objects.all(), start=2):
        for index, column in enumerate(columns):
            col_letter = get_column_letter(index + 1)
            ws[f"{col_letter}{row}"] = getattr(obj, column.column_name)

    file_obj = io.BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)

    return file_obj


def create_scheduled_task(instance: models.ScheduledReport) -> None:
    """
    Creates or updates a scheduled task for sending reports based on the schedule type provided in
    the instance object.

    Args:
        instance (ScheduledReport): An instance of the ScheduledReport model containing the schedule
        details.

    Returns:
        None

    Raises:
        exceptions.InvalidScheduleType: Raised when an invalid schedule type is provided.

    The function creates or updates a schedule using CrontabSchedule objects from Django. The
    schedule type is determined from the instance object, which can be daily, weekly, or monthly. If the
    schedule type is daily, a schedule is created or updated based on the instance time and timezone. If
    the schedule type is weekly, a schedule is created or updated based on the instance day of the week,
    time, and timezone. If the schedule type is monthly, a schedule is created or updated based on the
    instance day of the month, time, and timezone.

    Once the schedule is created or updated, a periodic task is created or updated using
    PeriodicTask objects from Django. The name of the task is created using the instance primary key.
    The schedule type and details are stored in the task, along with the function to call and the
    arguments to pass to the function. The function called is 'reports.tasks.send_scheduled_report', and
    the argument is a JSON-encoded string containing the primary key of the ScheduledReport instance.

    If a periodic task with the same name already exists, the task is updated with the new schedule
    and arguments. If the task type is not 'crontab', the task's 'interval' attribute is set to the new
    schedule.

    If an invalid schedule type is provided, a ValueError is raised.
    """
    if instance.schedule_type == models.ScheduleType.DAILY:
        schedule, _ = CrontabSchedule.objects.update_or_create(
            hour=instance.time.hour,
            minute=instance.time.minute,
            timezone=instance.timezone,
        )
        task_type = "crontab"
    elif instance.schedule_type == models.ScheduleType.WEEKLY:
        weekdays = ",".join([str(weekday.id) for weekday in instance.day_of_week.all()])
        schedule, _ = CrontabSchedule.objects.update_or_create(
            day_of_week=weekdays,
            hour=instance.time.hour,
            minute=instance.time.minute,
            timezone=instance.timezone,
        )
        task_type = "crontab"
    elif instance.schedule_type == models.ScheduleType.MONTHLY:
        schedule, _ = CrontabSchedule.objects.update_or_create(
            day_of_month=instance.day_of_month,
            hour=instance.time.hour,
            minute=instance.time.minute,
            timezone=instance.timezone,
        )
        task_type = "crontab"
    else:
        raise exceptions.InvalidScheduleTypeException("Invalid schedule type.")

    task, created_task = PeriodicTask.objects.update_or_create(
        name=f"Send scheduled report {instance.user_id}-{instance.pk}",
        defaults={
            "crontab": schedule if task_type == "crontab" else None,
            "interval": schedule if task_type == "interval" else None,
            "task": "reports.tasks.send_scheduled_report",
            "args": json.dumps([str(instance.pk)]),
        },
    )

    if not created_task:
        setattr(task, task_type, schedule)
        task.args = json.dumps([str(instance.pk)])
        task.save()
