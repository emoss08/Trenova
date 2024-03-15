# --------------------------------------------------------------------------------------------------
#  COPYRIGHT(c) 2024 Trenova                                                                       -
#                                                                                                  -
#  This file is part of Trenova.                                                                   -
#                                                                                                  -
#  The Trenova software is licensed under the Business Source License 1.1. You are granted the right
#  to copy, modify, and redistribute the software, but only for non-production use or with a total -
#  of less than three server instances. Starting from the Change Date (November 16, 2026), the     -
#  software will be made available under version 2 or later of the GNU General Public License.     -
#  If you use the software in violation of this license, your rights under the license will be     -
#  terminated automatically. The software is provided "as is," and the Licensor disclaims all      -
#  warranties and conditions. If you use this license's text or the "Business Source License" name -
#  and trademark, you must comply with the Licensor's covenants, which include specifying the      -
#  Change License as the GPL Version 2.0 or a compatible license, specifying an Additional Use     -
#  Grant, and not modifying the license in any other way.                                          -
# --------------------------------------------------------------------------------------------------

import io
import json

from django.apps import apps
from django.core.files import File
from django.core.mail import EmailMessage
from django.db.models import Model
from django.utils import timezone
from django_celery_beat.models import PeriodicTask
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reports import exceptions, helpers, models, selectors
from utils.types import ModelUUID


def get_model_by_table_name(table_name: str) -> type[Model] | None:
    """
    Returns a model class from Django apps by a given table name.
    If the model class is not found, it returns None.

    Args:
        table_name (str): A string representing the name of the database table.

    Returns:
        Type[models.Model]: A model class that corresponds to the given table name.
    """
    return next(
        (model for model in apps.get_models() if model._meta.db_table == table_name),
        None,
    )


def generate_excel_report_as_file(report: models.CustomReport) -> io.BytesIO:
    """Generate an Excel report as a file.

    Args:
        report (CustomReport): A Report object that specifies the table and columns to use for the report.

    Returns:
        io.BytesIO: A BytesIO object that contains the generated Excel file.

    This function takes a Report object and generates an Excel file based on the specified
    table and columns.
    The function uses openpyxl library to create a new workbook, writes the headers of the
    columns, and populates the data rows by iterating through the data model object.
    The generated Excel file is saved to a BytesIO object which is then returned to the caller.
    """
    model = get_model_by_table_name(report.table)

    if not model:
        raise exceptions.InvalidTableException("Invalid table name.")

    columns = report.columns.all().order_by("column_order")

    wb = Workbook()
    ws = wb.active

    for index, column in enumerate(columns):
        col_letter = get_column_letter(index + 1)
        ws[f"{col_letter}1"] = column.column_name

    for row, obj in enumerate(model.objects.all(), start=2):  # type: ignore
        for index, column in enumerate(columns):
            col_letter = get_column_letter(index + 1)
            ws[f"{col_letter}{row}"] = getattr(obj, column.column_name)

    file_obj = io.BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)

    return file_obj


def create_scheduled_task(*, instance: models.ScheduledReport) -> None:
    """Create or update a PeriodicTask for the given scheduled report instance.

    Args:
        instance (models.ScheduledReport): The scheduled report instance.

    Returns:
        None: This function does not return anything.

    The function first retrieves or creates the appropriate CrontabSchedule object
    based on the schedule type and scheduled report instance. Then, it creates or updates
    a PeriodicTask object with the necessary parameters.
    """
    schedule, task_type = selectors.get_crontab_schedule(
        schedule_type=instance.schedule_type, instance=instance
    )

    task, created_task = PeriodicTask.objects.update_or_create(
        name=f"Send scheduled report {instance.user_id}-{instance.pk}",
        defaults={
            "crontab": schedule if task_type == "crontab" else None,
            "interval": schedule if task_type == "interval" else None,
            "task": "send_scheduled_report",
            "kwargs": json.dumps({"report_id": str(instance.pk)}),
            "start_time": timezone.now(),
        },
    )

    if not created_task:
        setattr(task, task_type, schedule)
        task.kwargs = json.dumps({"report_id": str(instance.pk)})
        task.save()


# TODO(Wolfred): At some point we may want to batch reports to avoid sending too many emails at once.
def generate_report(
    *,
    model_name: str,
    columns: list[str],
    user_id: ModelUUID,
    file_format: str,
    delivery_method: str,
    email_recipients: list[str] | None,
) -> None:
    """Generate and deliver a report based on the specified parameters.

    This function validates the model, retrieves the user and model, prepares the data frame,
    generates the report file, and then delivers the report to the user based on the specified
    delivery method.

    Args:
        model_name (str): The name of the model to generate the report for.
        columns (list[str]): A list of column names to include in the report.
        user_id (ModelUUID): The unique identifier of the user for whom the report is generated.
        file_format (str): The format of the report file (e.g., 'csv', 'xlsx', 'pdf').
        delivery_method (str): The method of delivery for the report ('email' or 'download').
        email_recipients (list[str] | None): A list of email recipients if the delivery method is email.

    Returns:
        None: this function does not return anything.

    Raises:
        InvalidModelException: If the specified model is not allowed or does not exist.
        InvalidDeliveryMethodException: If the specified delivery method is not supported.

    Note:
        The function interacts with several helper functions and external models to accomplish its tasks,
        including validating the model, fetching user and model information, preparing the data frame,
        generating the report file, and handling the report delivery.
    """

    # Validate the delivery method.
    if delivery_method not in ["email", "local"]:
        raise exceptions.InvalidDeliveryMethodException("Invalid delivery method.")

    # Validate the model and get the user and model instances.
    helpers.validate_model(model_name=model_name)
    user, model = helpers.get_user_and_model(user_id=user_id, model_name=model_name)

    # Prepare the dataframe and generate the report file.
    df = helpers.prepare_dataframe(
        model=model, columns=columns, user=user, model_name=model_name
    )
    report_buffer, file_name = helpers.generate_report_file(
        df=df, model_name=model_name, file_format=file_format, user=user
    )

    # Save the report file and deliver it to the user.
    report_buffer.seek(0)
    django_file = File(report_buffer, name=file_name)

    new_report = models.UserReport.objects.create(
        organization=user.organization,
        user=user,
        report=django_file,
        business_unit=user.business_unit,
    )
    delivery_functions = {
        "email": lambda: helpers.delivery_email_report(
            buffer=report_buffer,
            user=user,
            email_recipients=email_recipients,
            file_name=file_name,
            model_name=model_name,
        ),
        "local": lambda: helpers.deliver_local_report(
            model_name=model_name, user=user, report_obj=new_report
        ),
    }

    delivery_func = delivery_functions.get(delivery_method)

    # If the delivery method is invalid, raise an exception.
    if not delivery_func:
        raise exceptions.InvalidDeliveryMethodException("Invalid delivery method.")

    delivery_func()


# TODO(Wolfred): Batch reports to avoid sending too many emails at once.
def generate_scheduled_report(*, report_id: str) -> None:
    """Generate a scheduled report and deliver it to the user who created it.

    The function retrieves the scheduled report instance, generates the report file,
    and delivers the report to the user who created it.

    Args:
        report_id (str): The unique identifier of the scheduled report.

    Returns:
        None: This function does not return anything.
    """

    scheduled_report = selectors.get_scheduled_report_by_id(report_id=report_id)

    if not scheduled_report.is_active:
        return

    report = scheduled_report.custom_report
    user = scheduled_report.user

    excel_file = generate_excel_report_as_file(report)

    email = EmailMessage(
        subject=f"Your scheduled report: {report.name}",
        body=f"Hi {user.profile.first_name},\n\nAttached is your scheduled report: {report.name}.",
        from_email="noreply@trenova.app",
        to=[user.email],
    )

    # TODO(Wolfred): Add support for multiple recipients and multiple file formats.
    email.attach(
        f"{report.name}.xlsx",
        excel_file.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    email.send()
    excel_file.close()
