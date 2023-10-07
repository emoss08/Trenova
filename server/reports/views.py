# --------------------------------------------------------------------------------------------------
#  COPYRIGHT(c) 2023 MONTA                                                                         -
#                                                                                                  -
#  This file is part of Monta.                                                                     -
#                                                                                                  -
#  The Monta software is licensed under the Business Source License 1.1. You are granted the right -
#  to copy, modify, and redistribute the software, but only for non-production use or with a total -
#  of less than three server instances. Starting from the Change Date (November 16, 2026), the     -
#  software will be made available under version 2 or later of the GNU General Public License.     -
#  If you use the software in violation of this license, your rights under the license will be     -
#  terminated automatically. The software is provided "as is," and the Licensor disclaims all      -
#  warranties and conditions. If you use this license's text or the "Business Source License" name -
#  and trademark, you must comply with the Licensor's covenants, which include specifying the      -
#  Change License as the GPL Version 2.0 or a compatible license, specifying an Additional Use     -
#  Grant, and not modifying the license in any other way.                                          -
# --------------------------------------------------------------------------------------------------

from django.http import HttpRequest, HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reports import models, services
from utils.models import generate_random_string


def generate_excel_report(request: HttpRequest, report_id: str) -> HttpResponse:
    """
    Generates an Excel report with data from a specified model class and columns.
    The report is based on a custom report object with a specified report_id.

    Args:
        request (HttpRequest): An object representing the request made to the server.
        report_id (str): A string representing the id of the custom report object.

    Returns:
        HttpResponse: An HTTP response with an Excel file containing the report data.
    """
    # Get the report and related columns
    report = models.CustomReport.objects.get(pk=report_id)
    columns = report.columns.all().order_by("column_shipment")

    # Get the model by table name
    model = services.get_model_by_table_name(report.table)
    if not model:
        return HttpResponse("Model not found", status=404)

    # Query the model to get the data
    queryset = model.objects.all()  # type: ignore

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = report.name

    # Write header row
    for index, column in enumerate(columns):
        col_letter = get_column_letter(index + 1)
        ws[f"{col_letter}1"] = column.column_name

    # Write data rows
    for row_num, obj in enumerate(queryset, start=2):
        for col_num, column in enumerate(columns):
            col_letter = get_column_letter(col_num + 1)
            value = getattr(obj, column.column_name)
            ws[f"{col_letter}{row_num}"] = value

    # Save the workbook to a temporary file and return it as a response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    random_string = generate_random_string()
    filename = f"{report.name}-{random_string}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
